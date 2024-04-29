import requests
from bs4 import BeautifulSoup
import openpyxl
from colorama import Fore


def buscar_productos(string):
    print('--- Búsqueda de Productos ---')
    try:
        r = requests.get('https://listado.mercadolibre.com.ar/{}#D[{}]'.format(string.replace(' ', '-'), string))
        contenido = r.content

        soup = BeautifulSoup(contenido, 'html.parser')
        alldivs = soup.find_all('div', {'class': 'andes-card'})

        products_array = []

        for item in alldivs:
            data = {}
            title_elem = item.find('h2', {'class': 'ui-search-item__title'})
            if title_elem:
                data['Articulo'] = title_elem.text
            else:
                data['Articulo'] = 'No disponible'

            price_elem = item.find('span', {'class': 'andes-money-amount__fraction'})
            if price_elem:
                data['Precio'] = price_elem.text
            else:
                data['Precio'] = 'No disponible'

            rebate_elem = item.find('span', {
                'class': 'andes-money-amount ui-search-price__part ui-search-price__part--medium andes-money-amount--cents-superscript'})
            if rebate_elem:
                data['Rebajado'] = rebate_elem.text
            else:
                data['Rebajado'] = 'No disponible'

            link_elem = item.find('a', {'class': 'ui-search-item__group__element'})
            if link_elem and 'href' in link_elem.attrs:
                data['link'] = link_elem['href']
            else:
                data['link'] = 'No disponible'

            products_array.append(data)
            # Mostrar datos en la consola con colores
            print(Fore.GREEN + f"Artículo: {data['Articulo']}")
            print(Fore.YELLOW + f"Precio: {data['Precio']}")
            print(Fore.CYAN + f"Rebajado: {data['Rebajado']}")
            print(Fore.BLUE + f"Enlace: {data['link']}")
            print("-" * 30)  # Separador entre productos

        if not products_array:
            print(Fore.YELLOW + 'No se encontraron productos para la búsqueda.')

        # Crear un nombre único para el archivo Excel
        excel_file = f'productos_{string.replace(" ", "_").lower()}.xlsx'

        # Verificar si el archivo existe y agregar un número único si es necesario
        counter = 1
        while any(excel_file in name for name in openpyxl.Workbook().sheetnames):
            excel_file = f'productos_{string.replace(" ", "_").lower()}_{counter}.xlsx'
            counter += 1

        # Crear un nuevo libro de Excel y una hoja de datos
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Productos'

        # Escribir los encabezados
        sheet.append(['Articulo', 'Precio', 'Rebajado', 'Link'])

        # Escribir los datos en la hoja de Excel
        for product in products_array:
            sheet.append([product['Articulo'], product['Precio'], product['Rebajado'], product['link']])

        # Guardar el libro de Excel
        workbook.save(excel_file)

        print(Fore.RESET + f"\nLos datos se han guardado en '{excel_file}'.")

        return products_array

    except Exception as e:
        print(Fore.RED + f"Error al procesar la búsqueda: {e}")
        return None


if __name__ == "__main__":
    string_busqueda = input('Ingrese el término que desea buscar en MercadoLibre: ')
    buscar_productos(string_busqueda)
